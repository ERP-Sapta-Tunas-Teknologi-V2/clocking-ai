from typing import Dict, List, Optional, Union, Tuple
import streamlit as st
import plotly.express as px
import requests
import json
import re
import mysql.connector
import pandas as pd
from fpdf import FPDF
import io
from decimal import Decimal
import plotly.io as pio
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as ReportLabImage
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from streamlit_modal import Modal
import os

# Load environment variables from a .env file if present
def load_env_file(env_path: str = ".env") -> None:
    if os.path.exists(env_path):
        try:
            with open(env_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    if "=" not in line:
                        continue
                    key, value = line.split("=", 1)
                    key = key.strip()
                    value = value.strip().strip('"').strip("'")
                    # Do not override existing environment variables
                    if key and key not in os.environ:
                        os.environ[key] = value
        except Exception:
            # Silently continue if .env cannot be read; fall back to defaults
            pass

# Ensure env vars are loaded before Config is defined
load_env_file()

# ================================
# ✅ CONFIGURATION
# ================================
class Config:
    OLLAMA_URL = "http://localhost:11434/api/generate"
    DB_CONFIG = {
        "host": os.getenv("DB_HOST", "localhost"),
        "user": os.getenv("DB_USER", "root"),
        "password": os.getenv("DB_PASSWORD", ""),
        "database": os.getenv("DB_NAME", "clocking_reports"),
    }
    MODEL_LIST = ["qwen3:0.6b"]
    SQL_MAPPING = {
        "sql1": {
            "description": "Jumlah clocking untuk user A dengan detail per category",
            "query": """
                SELECT 
                    cc.category_description,
                    COUNT(ca.activity_id) AS total_clocking,
                    SUM(ca.duration_minutes) AS total_minutes
                FROM clocking_activities ca
                JOIN daily_activities da ON ca.daily_activity_id = da.daily_activity_id
                JOIN users u ON da.user_id = u.user_id
                JOIN category_clocking cc ON ca.category_id = cc.category_id
                WHERE u.full_name LIKE %s
                GROUP BY cc.category_id, cc.category_description
                ORDER BY total_minutes DESC;
            """
        },
        "sql2": {
            "description": "Top 5 over clocking & Top 5 under clocking users dengan detail category",
            "query": """
                WITH WeeklyClocking AS (
                    SELECT 
                        u.full_name,
                        cc.category_description,
                        WEEK(ca.start_date) AS week_number,
                        SUM(ca.duration_minutes) / 60.0 AS total_hours
                    FROM clocking_activities ca
                    JOIN daily_activities da ON ca.daily_activity_id = da.daily_activity_id
                    JOIN users u ON da.user_id = u.user_id
                    JOIN category_clocking cc ON ca.category_id = cc.category_id
                    GROUP BY u.user_id, u.full_name, cc.category_id, cc.category_description, WEEK(ca.start_date)
                ),
                OverUnderClocking AS (
                    SELECT 
                        full_name,
                        category_description,
                        AVG(total_hours) AS avg_weekly_hours,
                        CASE 
                            WHEN AVG(total_hours) > 40 THEN 'Overclocking'
                            ELSE 'Underclocking'
                        END AS clocking_status
                    FROM WeeklyClocking
                    GROUP BY full_name, category_description
                )
                (SELECT full_name, category_description, avg_weekly_hours, clocking_status
                 FROM OverUnderClocking
                 WHERE clocking_status = 'Overclocking'
                 ORDER BY avg_weekly_hours DESC
                 LIMIT 5)
                UNION ALL
                (SELECT full_name, category_description, avg_weekly_hours, clocking_status
                 FROM OverUnderClocking
                 WHERE clocking_status = 'Underclocking'
                 ORDER BY avg_weekly_hours ASC
                 LIMIT 5);
            """
        },
        "sql3": {
            "description": "Analisa efisiensi user B pada bulan tertentu",
            "query": """
                SELECT 
                    u.full_name,
                    SUM(ca.duration_minutes) / 60.0 AS total_hours,
                    40 * 4 AS target_hours_month,
                    (SUM(ca.duration_minutes) / 60.0) - (40 * 4) AS difference_from_target,
                    CASE 
                        WHEN SUM(ca.duration_minutes) / 60.0 >= 40 * 4 THEN 'Efficient'
                        ELSE 'Not Efficient'
                    END AS efficiency_status
                FROM clocking_activities ca
                JOIN daily_activities da ON ca.daily_activity_id = da.daily_activity_id
                JOIN users u ON da.user_id = u.user_id
                WHERE u.full_name LIKE %s
                    AND MONTH(ca.start_date) = %s
                    AND YEAR(ca.start_date) = YEAR(CURDATE())
                GROUP BY u.user_id, u.full_name;
            """
        },
        "sql4": {
            "description": "Analisa clocking user dari bulan tertentu hingga bulan tertentu dibandingkan target",
            "query": """
                SELECT 
                    DATE_FORMAT(ca.start_date, '%Y-%m') AS month,
                    SUM(ca.duration_minutes) / 60.0 AS total_hours,
                    40 * 4 AS monthly_target_hours,
                    (SUM(ca.duration_minutes) / 60.0) - (40 * 4) AS difference_from_target
                FROM clocking_activities ca
                JOIN daily_activities da ON ca.daily_activity_id = da.daily_activity_id
                JOIN users u ON da.user_id = u.user_id
                WHERE u.full_name LIKE %s
                    AND MONTH(ca.start_date) BETWEEN %s AND %s
                    AND YEAR(ca.start_date) = YEAR(CURDATE())
                GROUP BY DATE_FORMAT(ca.start_date, '%Y-%m')
                ORDER BY month ASC;
            """
        },
        "sql5": {
            "description": "Grafik clocking Month Of Month selama 4 bulan untuk user D pada category 400",
            "query": """
                SELECT 
                    DATE_FORMAT(ca.start_date, '%Y-%m') AS month,
                    SUM(ca.duration_minutes) / 60.0 AS total_hours
                FROM clocking_activities ca
                JOIN daily_activities da ON ca.daily_activity_id = da.daily_activity_id
                JOIN users u ON da.user_id = u.user_id
                WHERE u.full_name LIKE %s
                    AND ca.category_id = 400
                    AND ca.start_date >= DATE_SUB(CURDATE(), INTERVAL 4 MONTH)
                GROUP BY DATE_FORMAT(ca.start_date, '%Y-%m')
                ORDER BY month ASC;
            """
        },
        "sql6": {
            "description": "Report clocking untuk Tim PM (top hingga under clocking)",
            "query": """
                WITH WeeklyClocking AS (
                    SELECT 
                        u.full_name,
                        p.project_name,
                        WEEK(ca.start_date) AS week_number,
                        SUM(ca.duration_minutes) / 60.0 AS total_hours
                    FROM clocking_activities ca
                    JOIN daily_activities da ON ca.daily_activity_id = da.daily_activity_id
                    JOIN users u ON da.user_id = u.user_id
                    JOIN project_users pu ON u.user_id = pu.user_id
                    JOIN projects p ON pu.project_code = p.project_code
                    WHERE p.project_manager_id IS NOT NULL
                    GROUP BY u.user_id, u.full_name, p.project_code, p.project_name, WEEK(ca.start_date)
                ),
                RankedClocking AS (
                    SELECT 
                        full_name,
                        project_name,
                        AVG(total_hours) AS avg_weekly_hours,
                        CASE 
                            WHEN AVG(total_hours) > 40 THEN 'Overclocking'
                            WHEN AVG(total_hours) < 40 THEN 'Underclocking'
                            ELSE 'On Target'
                        END AS clocking_status
                    FROM WeeklyClocking
                    GROUP BY full_name, project_name
                )
                SELECT 
                    full_name,
                    project_name,
                    avg_weekly_hours,
                    clocking_status
                FROM RankedClocking
                ORDER BY avg_weekly_hours DESC;
            """
        }
    }

# ================================
# ✅ DATABASE HANDLER
# ================================

class Database:
    @staticmethod
    def run_query(sql: str, params: tuple = None) -> Union[List[Dict], str]:
        try:
            with mysql.connector.connect(**Config.DB_CONFIG) as conn:
                with conn.cursor(dictionary=True) as cursor:
                    cursor.execute(sql, params)
                    return cursor.fetchall()
        except mysql.connector.Error as e:
            return f"Database error: {e}"

# ================================
# ✅ LLM HANDLER
# ================================
class LLM:
    # Define synonyms for key terms to match query variations
    KEYWORD_SYNONYMS = {
        "analisa": ["analisa", "analisis", "evaluasi", "review", "tinjau"],
        "user": ["user", "pengguna"],
        "bulan": ["bulan", "month", "months"],
        "clocking": ["clocking", "jam kerja", "waktu kerja", ""],  # Empty string makes clocking optional
        "4 bulan": ["4 bulan", "empat bulan", "4 months", "last 4 months"],
        "jumlah": ["jumlah", "total", "banyak"],
        "category": ["category", "kategori", "jenis"],
        "overclocking": ["overclocking", "over clocking", "kelebihan jam", "jam lebih"],
        "underclocking": ["underclocking", "under clocking", "kekurangan jam", "jam kurang"],
        "tim pm": ["tim pm", "tim project manager", "tim manajer proyek", "project manager team"],
        "efisien": ["efisien", "efisiensi", "produktif"],
        "grafik": ["grafik", "chart", "diagram"],
        "month_names": ["januari", "februari", "maret", "april", "mei", "juni", "juli", "agustus", "september", "oktober", "november", "desember"]
    }

    @staticmethod
    def extract_username(query: str) -> Optional[str]:
        # Extract username after user-related keywords
        match = re.search(r"(?:user|pengguna|untuk user|analisa user)\s+(\w+)", query, re.IGNORECASE)
        return match.group(1) if match else None

    @staticmethod
    def extract_month_range(query: str) -> Optional[Tuple[int, int]]:
        normalized_query = LLM.normalize_query(query)
        month_names = LLM.KEYWORD_SYNONYMS["month_names"]

        # Numeric range: bulan 2-9
        numeric_range_match = re.search(r"bulan\s+(\d+)-(\d+)", normalized_query, re.IGNORECASE)
        if numeric_range_match:
            start_month, end_month = int(numeric_range_match.group(1)), int(numeric_range_match.group(2))
            if 1 <= start_month <= 12 and 1 <= end_month <= 12 and start_month <= end_month:
                return start_month, end_month

        # Named month range: januari-maret
        name_range_match = re.search(r"\b(\w+)-(\w+)\b", normalized_query, re.IGNORECASE)
        if name_range_match:
            start_name, end_name = name_range_match.group(1).lower(), name_range_match.group(2).lower()
            if start_name in month_names and end_name in month_names:
                start_month = month_names.index(start_name) + 1
                end_month = month_names.index(end_name) + 1
                if start_month <= end_month:
                    return start_month, end_month

        # Single month numeric: bulan 3
        single_numeric_match = re.search(r"bulan\s+(\d+)", normalized_query, re.IGNORECASE)
        if single_numeric_match:
            month = int(single_numeric_match.group(1))
            if 1 <= month <= 12:
                return month, month

        # Single month name: maret
        month_pattern = '|'.join(month_names)
        single_name_match = re.search(rf"\b({month_pattern})\b", normalized_query, re.IGNORECASE)
        if single_name_match:
            month_name = single_name_match.group(1).lower()
            month = month_names.index(month_name) + 1
            return month, month

        return None

    @staticmethod
    def normalize_query(query: str) -> str:
        # Normalize query: lowercase, remove extra spaces
        return ' '.join(query.lower().strip().split())

    @classmethod
    def detect_sql_query_type(cls, query: str) -> Optional[Tuple[str, Optional[Tuple[int, int]]]]:
        normalized_query = cls.normalize_query(query)

        # Helper function to check if any synonym is present
        def has_synonym(keyword: str, text: str) -> bool:
            return any(syn in text for syn in cls.KEYWORD_SYNONYMS[keyword])

        # Extract month range for month-specific queries
        month_range = cls.extract_month_range(query)

        # sql3: Analisa efisiensi user pada bulan tertentu (efisien keyword is optional)
        if has_synonym("analisa", normalized_query) and has_synonym("user", normalized_query) and month_range and month_range[0] == month_range[1]:
            return "sql3", month_range

        # sql4: Analisa user dari bulan tertentu hingga bulan tertentu
        if has_synonym("analisa", normalized_query) and has_synonym("user", normalized_query) and month_range and month_range[0] != month_range[1]:
            return "sql4", month_range

        # sql1: Jumlah clocking untuk user A dengan detail per category
        if has_synonym("jumlah", normalized_query) and has_synonym("user", normalized_query) and has_synonym("category", normalized_query):
            return "sql1", None

        # sql2: Top 5 overclocking & Top 5 underclocking
        if has_synonym("overclocking", normalized_query) or has_synonym("underclocking", normalized_query):
            return "sql2", None

        # sql5: Grafik clocking Month Of Month selama 4 bulan untuk user D pada category 400
        if has_synonym("grafik", normalized_query) and has_synonym("4 bulan", normalized_query) and has_synonym("user", normalized_query):
            return "sql5", None

        # sql6: Report clocking untuk Tim PM
        if has_synonym("tim pm", normalized_query):
            return "sql6", None
        
        # If no SQL type matched
        return None

    @staticmethod
    def stream_response(model: str, prompt: str) -> Tuple[str, str]:
        response = requests.post(
            Config.OLLAMA_URL,
            json={"model": model, "prompt": prompt, "stream": True},
            stream=True
        )
        full_think, full_response, current_section = "", "", "response"
        
        for line in response.iter_lines():
            if not line:
                continue
            try:
                data = json.loads(line.decode('utf-8'))
                chunk = data.get('response', '')
                if "<think>" in chunk:
                    current_section = "think"
                    full_think += re.sub(r'^<think>', '', chunk)
                elif "</think>" in chunk:
                    current_section = "response"
                    full_think += re.sub(r'</think>$', '', chunk)
                elif current_section == "think":
                    full_think += chunk
                else:
                    full_response += chunk
            except json.JSONDecodeError:
                continue
        return full_response.strip() or "[No response from model]", full_think.strip() or ""

    @classmethod
    def summarize(cls, model: str, result: List[Dict], query: str) -> Tuple[str, str]:
        prompt = f"Berikut hasil query:\n{json.dumps(result, default=str)}\n\nTolong buatkan ringkasan berdasarkan query ini: {query}"
        return cls.stream_response(model, prompt)

# ================================
# ✅ OUTPUT GENERATOR
# ================================

class OutputGenerator:
    @staticmethod
    def to_excel(data: List[Dict], query: str, think: str, response: str, chart_image: Optional[bytes] = None) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Analysis Result"

        # Write query
        ws['A1'] = "Query"
        ws['B1'] = query
        ws['A1'].font = ws['A1'].font.copy(bold=True)
        
        # Write data
        if data:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, start=1):
                ws.cell(row=3, column=col).value = header
                ws.cell(row=3, column=col).font = ws.cell(row=3, column=col).font.copy(bold=True)
            
            for row, item in enumerate(data, start=4):
                for col, key in enumerate(headers, start=1):
                    # Format month as string if it's a datetime
                    value = item[key]
                    if key == 'month' and isinstance(value, pd.Timestamp):
                        value = value.strftime('%b %Y')
                    ws.cell(row=row, column=col).value = value
        
        # Write thinking process
        ws['A' + str(ws.max_row + 2)] = "Thinking Process"
        ws['B' + str(ws.max_row)] = think
        ws['A' + str(ws.max_row)].font = ws['A' + str(ws.max_row)].font.copy(bold=True)
        
        # Write response
        ws['A' + str(ws.max_row + 2)] = "Response"
        ws['B' + str(ws.max_row)] = response
        ws['A' + str(ws.max_row)].font = ws['A' + str(ws.max_row)].font.copy(bold=True)

        # Add chart image if provided
        if chart_image:
            img_stream = BytesIO(chart_image)
            img = Image(img_stream)
            ws.add_image(img, f'A{ws.max_row + 2}')
            ws.row_dimensions[ws.max_row + 1].height = 300  # Adjust row height for image

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def to_pdf(query: str, data: List[Dict], analysis: str, chart_image: Optional[bytes] = None) -> bytes:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []

        # Add query
        story.append(Paragraph("Query:", styles['Heading2']))
        story.append(Paragraph(query, styles['Normal']))
        story.append(Spacer(1, 12))

        # Add data
        if data:
            story.append(Paragraph("Data:", styles['Heading2']))
            for item in data:
                # Format month as string if it's a datetime
                formatted_item = {k: v.strftime('%b %Y') if k == 'month' and isinstance(v, pd.Timestamp) else v for k, v in item.items()}
                data_str = ", ".join([f"{key}: {value}" for key, value in formatted_item.items()])
                story.append(Paragraph(data_str, styles['Normal']))
            story.append(Spacer(1, 12))

        # Add analysis
        story.append(Paragraph("Analysis:", styles['Heading2']))
        story.append(Paragraph(analysis.replace('\n', '<br/>'), styles['Normal']))
        story.append(Spacer(1, 12))

        # Add chart image if provided
        if chart_image:
            story.append(Paragraph("Chart:", styles['Heading2']))
            img_stream = BytesIO(chart_image)
            img = ReportLabImage(img_stream, width=500, height=300)
            story.append(img)

        doc.build(story)
        return buffer.getvalue()

# ================================
# ✅ STREAMLIT UI
# ================================

def main():
    st.title("📊 LLM + MySQL Clocking Analysis")
    st.sidebar.title("📁 Query History")
    
    # Initialize session state
    if 'history' not in st.session_state:
        st.session_state.history = []
    if 'last_query' not in st.session_state:
        st.session_state.last_query = ""
    if 'last_result' not in st.session_state:
        st.session_state.last_result = None
    if 'last_response' not in st.session_state:
        st.session_state.last_response = ""
    if 'last_think' not in st.session_state:
        st.session_state.last_think = ""
    if 'last_chart_image' not in st.session_state:
        st.session_state.last_chart_image = None
    if 'last_sql_id' not in st.session_state:
        st.session_state.last_sql_id = None
    if 'last_username' not in st.session_state:
        st.session_state.last_username = None
    if 'selected_history_index' not in st.session_state:
        st.session_state.selected_history_index = None
    if 'modal_open' not in st.session_state:
        st.session_state.modal_open = False

    # Initialize modal
    modal = Modal("Report", key="report_modal", padding=20, max_width=800)

    # Display history in sidebar immediately
    for i, entry in enumerate(st.session_state.history):
        with st.sidebar.expander(f"Q{i+1}: {entry['query'][:30]}..."):
            st.write(f"**Query:** {entry['query'][:30]}...")
            st.write(f"**Answer:** {entry['response'][:30]}...")
            if st.button("View Report", key=f"view_report_{i}"):
                st.session_state.selected_history_index = i
                st.session_state.modal_open = True
                st.rerun()  # Force rerun to open modal

    # Open modal if triggered
    if st.session_state.modal_open and st.session_state.selected_history_index is not None:
        entry = st.session_state.history[st.session_state.selected_history_index]
        if modal.open():
            st.write(f"**Query:** {entry['query']}")
            if entry['result']:
                st.write("**Result:**")
                st.json(entry['result'])
            if entry['sql_id'] == "sql5" and entry['result']:
                df = pd.DataFrame(entry['result'])
                if 'month' in df.columns and 'total_hours' in df.columns:
                    df['month'] = pd.to_datetime(df['month'], format='%Y-%m').dt.strftime('%b %Y')
                    df['total_hours'] = df['total_hours'].astype(float)
                    fig = px.line(
                        df,
                        x='month',
                        y='total_hours',
                        title=f"Clocking Hours for User {entry['username']} (Category 400, Last 4 Months)",
                        labels={'month': 'Month', 'total_hours': 'Total Hours'},
                        markers=True
                    )
                    fig.update_layout(
                        xaxis_title="Month",
                        yaxis_title="Total Hours",
                        showlegend=False
                    )
                    st.plotly_chart(fig, key=f"sql5_chart_history_{st.session_state.selected_history_index}")
            st.markdown(f"**Summary:** {entry['response']}")
            if entry['think']:
                with st.expander("🧠 Thinking Process", expanded=False):
                    st.markdown(entry['think'])
            # Download buttons for selected history entry
            if entry['result']:
                st.download_button(
                    label="📥 Download Excel (with Analysis)",
                    data=OutputGenerator.to_excel(
                        entry['result'],
                        entry['query'],
                        entry['think'],
                        entry['response'],
                        entry['chart_image']
                    ),
                    file_name=f"analysis_output_Q{st.session_state.selected_history_index + 1}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"excel_download_{st.session_state.selected_history_index}"
                )
                st.download_button(
                    label="📄 Download PDF Report",
                    data=OutputGenerator.to_pdf(
                        entry['query'],
                        entry['result'],
                        f"[THINKING]\n{entry['think']}\n\n[RESPONSE]\n{entry['response']}",
                        entry['chart_image']
                    ),
                    file_name=f"analysis_report_Q{st.session_state.selected_history_index + 1}.pdf",
                    mime="application/pdf",
                    key=f"pdf_download_{st.session_state.selected_history_index}"
                )
            st.session_state.modal_open = False  # Close modal after interaction

    selected_model = st.selectbox("Select Model", Config.MODEL_LIST, index=0)
    query = st.text_area("Enter your query:", value=st.session_state.last_query, placeholder="e.g. clocking Month Of Month selama 4 bulan untuk user juanrico.")
    submit_button = st.button("Submit")

    if submit_button and query:
        # Store query in session state
        st.session_state.last_query = query
        sql_result = LLM.detect_sql_query_type(query)
        chart_image = None

        if sql_result:
            sql_id, month_range = sql_result
            st.session_state.last_sql_id = sql_id

            # Skip username check for sql2 and sql6
            if sql_id not in ["sql2", "sql6"]:
                username = LLM.extract_username(query)
                if not username:
                    st.error("❌ Username not found in query.")
                    return
                st.session_state.last_username = username
                if sql_id == "sql3":
                    query_params = (f"%{username}%", month_range[0])
                elif sql_id == "sql4":
                    query_params = (f"%{username}%", month_range[0], month_range[1])
                else:
                    query_params = (f"%{username}%",)
            else:
                query_params = None
                st.session_state.last_username = None

            result = Database.run_query(Config.SQL_MAPPING[sql_id]["query"], query_params)
            if isinstance(result, str):
                st.error(result)
                return

            st.session_state.last_result = result
            st.success("✅ SQL Executed Successfully")
            st.write("📊 Result:")
            st.json(result)

            # Generate Plotly chart for sql5
            if sql_id == "sql5" and result:
                df = pd.DataFrame(result)
                if 'month' in df.columns and 'total_hours' in df.columns:
                    df['month'] = pd.to_datetime(df['month'], format='%Y-%m').dt.strftime('%b %Y')
                    df['total_hours'] = df['total_hours'].astype(float)
                    fig = px.line(
                        df,
                        x='month',
                        y='total_hours',
                        title=f"Clocking Hours for User {username} (Category 400, Last 4 Months)",
                        labels={'month': 'Month', 'total_hours': 'Total Hours'},
                        markers=True
                    )
                    fig.update_layout(
                        xaxis_title="Month",
                        yaxis_title="Total Hours",
                        showlegend=False
                    )
                    st.plotly_chart(fig, key="sql5_chart_submit")
                    chart_image_buffer = BytesIO()
                    pio.write_image(fig, file=chart_image_buffer, format='png')
                    chart_image = chart_image_buffer.getvalue()
                    st.session_state.last_chart_image = chart_image
                else:
                    st.warning("⚠️ Unable to generate chart: Invalid data format.")

            st.info("🤖 Sending to LLM for analysis...")
            response, think = LLM.summarize(selected_model, result, query)
            st.session_state.last_response = response
            st.session_state.last_think = think

            # Store in history
            st.session_state.history.append({
                "query": query,
                "response": response,
                "think": think,
                "result": result,
                "sql_id": sql_id,
                "username": st.session_state.last_username,
                "chart_image": chart_image
            })
            st.session_state.selected_history_index = None  # Reset to show latest query by default

        else:
            st.info("🧠 No SQL matched. Sending directly to LLM...")
            response_container = st.empty()
            think_container = st.empty()
            try:
                response, think = LLM.stream_response(selected_model, query)
                if think:
                    with think_container.expander("🧠 Thinking Process", expanded=False):
                        st.markdown(think)
                if response:
                    response_container.markdown(f"**Response:** {response}")
                    st.session_state.last_response = response
                    st.session_state.last_think = think
                    st.session_state.last_query = query
                    st.session_state.history.append({
                        "query": query,
                        "response": response,
                        "think": think,
                        "result": None,
                        "sql_id": None,
                        "username": None,
                        "chart_image": None
                    })
                    st.session_state.selected_history_index = None
            except Exception as e:
                st.error(f"❌ Error: {e}")

    # Display LLM summary only for latest query
    if st.session_state.last_response:
        st.success("✅ LLM Executed Successfully")
        st.markdown(f"**Summary:** {st.session_state.last_response}")
        if st.session_state.last_think:
            with st.expander("🧠 Thinking Process", expanded=False):
                st.markdown(st.session_state.last_think)

        # Download buttons for last query
        if st.session_state.last_result:
            st.download_button(
                label="📥 Download Excel (with Analysis)",
                data=OutputGenerator.to_excel(
                    st.session_state.last_result,
                    st.session_state.last_query,
                    st.session_state.last_think,
                    st.session_state.last_response,
                    st.session_state.last_chart_image
                ),
                file_name="analysis_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="excel_download_latest"
            )
            st.download_button(
                label="📄 Download PDF Report",
                data=OutputGenerator.to_pdf(
                    st.session_state.last_query,
                    st.session_state.last_result,
                    f"[THINKING]\n{st.session_state.last_think}\n\n[RESPONSE]\n{st.session_state.last_response}",
                    st.session_state.last_chart_image
                ),
                file_name="analysis_report.pdf",
                mime="application/pdf",
                key="pdf_download_latest"
            )

if __name__ == "__main__":
    main()